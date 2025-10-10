import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

def to_get_news():
    # 百度热搜榜
    url = 'http://top.baidu.com/buzz?b=1&fr=topindex'
    # 构造请求头
    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.82 Safari/537.36'
    }
    try:
        r = requests.get(url, headers=headers, timeout=30)
        r.raise_for_status()
        r.encoding = r.apparent_encoding
        soup = BeautifulSoup(r.text, 'html.parser')

        title_list = soup.find_all(attrs={'class': 'c-single-text-ellipsis'})
        hot_list = soup.find_all(attrs={'class': 'hot-index_1Bl1a'})

        return title_list, hot_list

    except Exception as e:
        print(f"Error fetching data: {e}")
        return None, None

def crea_xlsx(title_list):
    if not title_list:
        print("Title list is empty. Cannot create Excel file.")
        return

    try:
        # 创建一个Workbook对象
        workbook = Workbook()
        # 激活 worksheet
        worksheet = workbook.active
        # 设置第一个单元格的值为'热搜榜'
        worksheet.cell(row=1, column=1, value='热搜榜')

        # 写入数据
        for idx, title in enumerate(title_list, start=2):
            worksheet.cell(row=idx, column=1, value=title.get_text())

        # 保存文件
        workbook.save('hot_list.xls')
        print("Excel file created successfully.")

    except Exception as e:
        print(f"Error creating Excel file: {e}")

if __name__ == '__main__':
    title_list, _ = to_get_news()
    crea_xlsx(title_list[:10])
